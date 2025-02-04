import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def crear_excel(datos_facturas, datos_proveedores, ruta_excel):
    """
    Crea un archivo Excel con:
    - 'FF' para datos extraídos de XML.
    - 'FX' para proveedores y valores extraídos de TXT.
    """
    try:
        if not datos_facturas and not datos_proveedores:
            print("No hay datos para exportar.")
            return

        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            # Hoja FF (Facturas XML)
            if datos_facturas:
                df_facturas = pd.DataFrame(datos_facturas)
                df_facturas.to_excel(writer, sheet_name="FF", index=False)
            
            # Hoja FX (Proveedores TXT)
            columnas_fx = ["ID", "Nombre proveedor", "TAX ID", "Invoice", "Fecha", "Moneda", "Total"]
            df_fx = pd.DataFrame(datos_proveedores, columns=columnas_fx) if datos_proveedores else pd.DataFrame(columns=columnas_fx)
            df_fx.to_excel(writer, sheet_name="FX", index=False)

        # Formato en FX
        wb = load_workbook(ruta_excel)
        ws_fx = wb["FX"]

        for cell in ws_fx[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for col in ws_fx.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws_fx.column_dimensions[col_letter].width = max_length + 2

        wb.save(ruta_excel)
        print(f"Datos exportados y formato aplicado en {ruta_excel}")

    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
